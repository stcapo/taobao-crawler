# 代码说明：
'''
代码功能： 基于ChromeDriver爬取taobao（淘宝）平台商品列表数据
输入参数:  KEYWORLD --> 搜索商品“关键词”；
          pageStart --> 爬取起始页；
          pageEnd --> 爬取终止页；
输出文件：爬取商品列表数据
        'Page'        ：页码
        'Num'         ：序号
        'title'       ：商品标题
        'Price'       ：商品价格
        'Deal'        ：商品销量
        'Location'    ：地理位置
        'Shop'        ：商品
        'IsPostFree'  ：是否包邮
        'Title_URL'   ：商品详细页链接
        'Shop_URL'    ：商铺链接
        'Img_URL'     ：图片链接
'''
# 声明第三方库/头文件
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time
import openpyxl as op  # 导入Excel读写库

# 全局变量
count = 1  # 写入Excel商品计数

KEYWORD = input('输入搜索的商品关键词Keyword：')  # 要搜索的商品的关键词
pageStart = int(input('输入爬取的起始页PageStart：'))  # 爬取起始页
pageEnd = int(input('输入爬取的终止页PageEnd：'))  # 爬取终止页

# 启动ChromeDriver服务
options = webdriver.ChromeOptions()
# 关闭自动测试状态显示 // 会导致浏览器报：请停用开发者模式
options.add_experimental_option("excludeSwitches", ['enable-automation'])
# 把chrome设为selenium驱动的浏览器代理；
driver = webdriver.Chrome(options=options)
# 反爬机制
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                       {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
# 窗口最大化
driver.maximize_window()
driver.get('https://www.taobao.com')
# wait是Selenium中的一个等待类，用于在特定条件满足之前等待一定的时间(这里是20秒)。
# 如果一直到等待时间都没满足则会捕获TimeoutException异常
wait = WebDriverWait(driver, 20)


# 输入“关键词”，搜索
def search_goods():
    try:
        print("正在搜索: {}".format(KEYWORD))
        # 找到搜索“输入框”
        input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#q")))
        # 找到“搜索”按钮
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#J_TSearchForm > div.search-button > button')))
        # 输入框写入“关键词KeyWord”
        input.send_keys(KEYWORD)
        # 点击“搜索”按键
        submit.click()
        # 搜索商品后会再强制停止2秒，如有滑块请手动操作
        time.sleep(2)
        print("搜索完成！")
    except Exception as exc:
        print("search_goods函数错误！Error：{}".format(exc))


# 翻页至第pageStar页
def turn_pageStart():
    try:
        print("正在翻转:第{}页".format(pageStart))
        # 滑动到页面底端
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # 滑动到底部后停留3s
        time.sleep(3)
        # 找到输入“页面”的表单，输入“起始页”
        pageInput = wait.until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/span[3]/input')))
        pageInput.send_keys(pageStart)
        # 找到页面跳转的“确定”按钮，并且点击
        admit = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/button[3]')))
        admit.click()
        print("已翻至:第{}页".format(pageStart))
    except Exception as exc:
        print("turn_pageStart函数错误！Error：{}".format(exc))


# 获取每一页的商品信息；
def get_goods(page):
    try:
        # 声明全局变量count
        global count
        # 刷新/滑动界面，使所有信息都加载完成后，按下Enter，开始爬取
        if input('确认界面加载完毕后，按下“Enter”开始爬取'):
            pass
        # 获取html网页
        html = driver.page_source
        doc = pq(html)
        # 提取所有商品的共同父元素的类选择器
        items = list(doc('div.content--CUnfXXxv > div > div').items())
        # print(items)
        for item in items:
            if item.find('.title--RoseSo8H').text() == '大家都在搜':
                pass
            else:
                # 定位商品标题
                title = item.find('.title--qJ7Xg_90 span').text()
                # 定位价格
                # price_int = item.find('.priceInt--yqqZMJ5a').text()
                # price_float = item.find('.priceFloat--XpixvyQ1').text()
                # if price_int and price_float:
                #     # price = float(f"{price_int}{price_float}")
                #     # price = price_int+price_float
                # else:
                #     price = 0
                price = item.find('.innerPriceWrapper--aAJhHXD4').text()
                price = float(price.replace('\n', '').replace('\r', ''))
                # 定位交易量
                deal = item.find('.realSales--XZJiepmt').text()
                deal = deal.replace("万", "0000")  # “万”字替换为0000
                deal = deal.split("人")[0]  # 以“人”分隔
                deal = deal.split("+")[0]  # 以“+”分隔
                # 定位所在地信息
                location = item.find('.procity--wlcT2xH9 span').text()
                # 定位店名
                shop = item.find('.shopNameText--DmtlsDKm').text()
                # 定位包邮的位置
                postText = item.find('.subIconWrapper--Vl8zAdQn').text()
                postText = "包邮" if "包邮" in postText else "/"
                # 定位商品url
                t_url = item.find('.doubleCardWrapperAdapt--mEcC7olq')
                t_url = t_url.attr('href')
                # t_url = item.attr('a.doubleCardWrapperAdapt--mEcC7olq href')
                # 定位店名url
                shop_url = item.find('.TextAndPic--grkZAtsC a')
                shop_url = shop_url.attr('href')
                # 定位商品图片url
                img = item.find('.mainPicAdaptWrapper--V_ayd2hD img')
                img_url = img.attr('src')
                # 构建商品信息字典
                product = {
                    'Page': page,
                    'Num': count - 1,
                    'title': title,
                    'price': price,
                    'deal': int(deal),
                    'location': location,
                    'shop': shop,
                    'isPostFree': postText,
                    'url': t_url,
                    'shop_url': shop_url,
                    'img_url': img_url
                }
                print(product)
                # 商品信息写入Excel表格中
                wb.cell(row=count, column=1, value=count - 1)  # 序号
                wb.cell(row=count, column=2, value=title)  # 标题
                wb.cell(row=count, column=3, value=price)  # 价格
                wb.cell(row=count, column=4, value=int(deal))  # 付款人数
                wb.cell(row=count, column=5, value=location)  # 地理位置
                wb.cell(row=count, column=6, value=shop)  # 店铺名称
                wb.cell(row=count, column=7, value=postText)  # 是否包邮
                wb.cell(row=count, column=8, value=t_url)  # 商品链接
                wb.cell(row=count, column=9, value=shop_url)  # 商铺链接
                wb.cell(row=count, column=10, value=img_url)  # 图片链接
                count += 1  # 下一行
    except Exception as exc:
        print("get_goods函数错误！Error：{}".format(exc))


# 翻页函数
def page_turning(page_number):
    try:
        print("正在翻页: 第{}页".format(page_number))
        # 强制等待2秒后翻页
        time.sleep(2)
        # 找到“下一页”的按钮
        submit = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/button[2]')))
        submit.click()
        # 判断页数是否相等
        wait.until(EC.text_to_be_present_in_element(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/span[1]/em'), str(page_number)))
        print("已翻至: 第{}页".format(page_number))
    except Exception as exc:
        print("page_turning函数错误！Error：{}".format(exc))


# 爬虫main函数
def Crawer_main():
    try:
        # 搜索KEYWORD
        search_goods()
        # 判断pageStart是否为第1页
        if pageStart != 1:
            turn_pageStart()
        # 爬取PageStart的商品信息
        get_goods(pageStart)
        # 从PageStart+1爬取到PageEnd
        for i in range(pageStart + 1, pageEnd + 1):
            page_turning(i)
            get_goods(i)
    except Exception as exc:
        print("Crawer_main函数错误！Error：{}".format(exc))


if __name__ == '__main__':
    # 建立Excel表格
    try:
        ws = op.Workbook()  # 创建Workbook
        wb = ws.create_sheet(index=0)  # 创建worsheet
        # Excel第一行：表头
        title_list = ['Num', 'title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree', 'Title_URL',
                      'Shop_URL', 'Img_URL']
        for i in range(0, len(title_list)):
            wb.cell(row=count, column=i + 1, value=title_list[i])
        count += 1  # 从第二行开始写爬取数据
    except Exception as exc:
        print("Excel建立失败！Error：{}".format(exc))

    # 开始爬取数据
    Crawer_main()

    # 保存Excel表格
    data = time.strftime('%Y%m%d-%H%M', time.localtime(time.time()))
    Filename = "{}_{}_FromTB.xlsx".format(KEYWORD, data)
    ws.save(filename=Filename)
    print(Filename + "存储成功~")